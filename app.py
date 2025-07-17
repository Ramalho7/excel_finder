import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox
import os

ctk.set_appearance_mode('dark')

folder_path = ""
search_term = ""

def upFolder():
    global folder_path 
    folder_path = filedialog.askdirectory(title="Selecione a pasta que deseja")

def search_item():
    global folder_path
    
    if not folder_path:
        messagebox.showwarning("Aviso", "Por favor, selecione uma pasta antes de buscar!")
        return
    
    search_term = search_entry.get()
    
    if not search_term.strip():
        messagebox.showwarning('Aviso', 'É preciso informar um item para buscar!')
        return
    
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
            file_path = os.path.join(folder_path, file_name)
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value and str(cell.value).strip().lower() == search_term.strip().lower():
                                results_box.insert(ctk.END, f"Encontrado: {cell.value} | Arquivo: {file_name} | Planilha: {sheet} | Célula: {cell.coordinate}\n")
                                os.startfile(file_path)
                                return
            except Exception as e:
                results_box.insert(ctk.END, f"Erro ao abrir {file_name}: {e}\n")
    results_box.insert(ctk.END, "Nenhum resultado encontrado.\n")


app = ctk.CTk()
app.title('Buscar item na planilha')

app.geometry('600x600')

search_label = ctk.CTkLabel(app, text="Item para buscar:")
search_label.pack(pady=10)

search_entry = ctk.CTkEntry(app, width=400)
search_entry.pack(pady=10)

search_button = ctk.CTkButton(app, text="Selecione a pasta", command=upFolder)
search_button.pack(pady=10)

app.bind('<Return>', lambda event: search_item())

search_button = ctk.CTkButton(app, text="Buscar", command=search_item)
search_button.pack(pady=10)

results_box = ctk.CTkTextbox(app, width=550, height=400)
results_box.pack(pady=10)

app.mainloop()